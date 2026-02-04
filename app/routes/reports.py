from flask import Blueprint, render_template, jsonify, send_file, request
from flask_login import login_required
from app import db
from app.models.models import Product
from app.models.transactions import Sale
from datetime import datetime, timedelta
from sqlalchemy import func, desc
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

bp = Blueprint('reports', __name__, url_prefix='/reports')

@bp.route('/')
@login_required
def index():
    """Página principal de reportes"""
    
    # Obtener fechas de filtro
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    
    if start_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
    else:
        start_date = datetime.now() - timedelta(days=30)
    
    if end_date_str:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
    else:
        end_date = datetime.now()
    
    # Ventas del período
    sales = Sale.query.filter(
        Sale.created_at >= start_date,
        Sale.created_at <= end_date
    ).order_by(Sale.created_at.desc()).all()
    
    # Debug: Verificar si hay ventas en total
    total_sales_count = Sale.query.count()
    print(f"DEBUG - Total ventas en BD: {total_sales_count}")
    print(f"DEBUG - Ventas en período ({start_date} a {end_date}): {len(sales)}")
    if sales:
        print(f"DEBUG - Primera venta: {sales[0].created_at}")
    
    # Ventas por categoría
    category_sales = db.session.query(
        Product.category,
        func.sum(Sale.total).label('total')
    ).join(
        Sale.items
    ).join(
        Product
    ).filter(
        Sale.created_at >= start_date,
        Sale.created_at <= end_date
    ).group_by(Product.category).all()
    
    category_sales = [
        {'category': cat, 'total': float(total)} 
        for cat, total in category_sales
    ]
    
    # Tendencia mensual (últimos 6 meses)
    monthly_trends = []
    for i in range(6, 0, -1):
        month_date = datetime.now() - timedelta(days=30*i)
        month_start = month_date.replace(day=1, hour=0, minute=0, second=0)
        if i > 1:
            next_month = month_date.replace(day=28) + timedelta(days=4)
            month_end = next_month.replace(day=1) - timedelta(days=1)
        else:
            month_end = datetime.now()
        
        month_sales = Sale.query.filter(
            Sale.created_at >= month_start,
            Sale.created_at <= month_end
        ).all()
        
        monthly_trends.append({
            'month': month_start.strftime('%b %Y'),
            'total': sum(s.total for s in month_sales)
        })
    
    # Top productos
    from app.models.transactions import SaleItem
    
    top_products_data = db.session.query(
        Product,
        func.sum(SaleItem.quantity).label('total_sold')
    ).join(
        SaleItem
    ).join(
        Sale
    ).filter(
        Sale.created_at >= start_date,
        Sale.created_at <= end_date
    ).group_by(Product.id).order_by(desc('total_sold')).limit(10).all()
    
    top_products = []
    for product, total_sold in top_products_data:
        product_dict = product.to_dict()
        product_dict['total_sold'] = int(total_sold)
        top_products.append(product_dict)
    
    # Productos con stock bajo
    low_stock = Product.query.filter(
        Product.stock <= Product.min_stock,
        Product.active == True
    ).all()
    
    # Convertir a diccionarios
    low_stock_dict = [p.to_dict() for p in low_stock]
    
    return render_template('reports/index.html',
                         sales=sales,
                         category_sales=category_sales,
                         monthly_trends=monthly_trends,
                         top_products=top_products,
                         low_stock=low_stock_dict,
                         start_date=start_date.strftime('%Y-%m-%d'),
                         end_date=end_date.strftime('%Y-%m-%d'))

@bp.route('/api/reports/dashboard')
@login_required
def dashboard_data():
    """API: Datos para el dashboard de reportes"""
    # Período (últimos 30 días)
    start_date = datetime.now() - timedelta(days=30)
    
    # Ventas del período
    sales = Sale.query.filter(Sale.created_at >= start_date).all()
    total_sales = sum(sale.total for sale in sales)
    total_profit = sum(
        sum((item.unit_price - item.product.cost_price) * item.quantity 
            for item in sale.items) 
        for sale in sales
    )
    
    # Ventas por categoría
    categories = {}
    for sale in sales:
        for item in sale.items:
            cat = item.product.category
            if cat not in categories:
                categories[cat] = 0
            categories[cat] += item.line_total
    
    # Top productos
    product_sales = {}
    for sale in sales:
        for item in sale.items:
            if item.product_id not in product_sales:
                product_sales[item.product_id] = {
                    'name': item.product_name,
                    'quantity': 0,
                    'total': 0
                }
            product_sales[item.product_id]['quantity'] += item.quantity
            product_sales[item.product_id]['total'] += item.line_total
    
    top_products = sorted(
        product_sales.values(), 
        key=lambda x: x['total'], 
        reverse=True
    )[:10]
    
    # Ventas diarias
    daily_sales = {}
    for sale in sales:
        date_key = sale.created_at.strftime('%Y-%m-%d')
        if date_key not in daily_sales:
            daily_sales[date_key] = 0
        daily_sales[date_key] += sale.total
    
    return jsonify({
        'total_sales': round(total_sales, 2),
        'total_profit': round(total_profit, 2),
        'sales_count': len(sales),
        'categories': categories,
        'top_products': top_products,
        'daily_sales': daily_sales
    })

@bp.route('/api/reports/export/excel')
@login_required
def export_excel():
    """API: Exportar reporte profesional de inventario a Excel"""
    wb = Workbook()
    
    # Colores por categoría (paleta profesional)
    CATEGORY_COLORS = {
        'Cuadernos': 'E3F2FD',     # Azul claro
        'Lápices': 'F3E5F5',      # Morado claro
        'Bolígrafos': 'E8F5E8',   # Verde claro
        'Marcadores': 'FFF3E0',   # Naranja claro
        'Colores': 'FCE4EC',      # Rosa claro
        'Tijeras': 'F1F8E9',      # Verde lima claro
        'Correctores': 'FFF8E1',  # Amarillo claro
        'Reglas': 'E0F2F1',       # Teal claro
        'Folders': 'F9FBE7',      # Lima claro
        'Mochilas': 'EDE7F6',     # Violeta claro
        'Calculadoras': 'E1F5FE', # Cian claro
        'Papel': 'FFF9C4',        # Ámbar claro
        'Arte': 'FFEBEE',         # Rojo claro
        'Accesorios': 'F5F5F5'    # Gris claro
    }
    
    # Estilos profesionales
    title_font = Font(name='Calibri', size=18, bold=True, color='1565C0')
    subtitle_font = Font(name='Calibri', size=14, bold=True, color='424242')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    data_font = Font(name='Calibri', size=10)
    
    header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    low_stock_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    no_stock_fill = PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")
    
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ============ HOJA 1: RESUMEN EJECUTIVO ============
    ws_resumen = wb.active
    ws_resumen.title = "📊 Resumen Ejecutivo"
    
    # Título principal
    ws_resumen['A1'] = "📦 LIBRERÍA JODA - REPORTE DE INVENTARIO PROFESIONAL"
    ws_resumen['A1'].font = title_font
    ws_resumen.merge_cells('A1:H1')
    ws_resumen['A1'].alignment = Alignment(horizontal='center')
    
    # Fecha de generación
    ws_resumen['A2'] = f"📅 Generado el: {datetime.now().strftime('%d de %B de %Y a las %H:%M hrs')}"
    ws_resumen['A2'].font = Font(name='Calibri', size=11, italic=True, color='666666')
    ws_resumen.merge_cells('A2:H2')
    ws_resumen['A2'].alignment = Alignment(horizontal='center')
    
    # Obtener datos de productos REALES del inventario
    products = Product.query.filter_by(active=True).all()
    
    # Si no hay productos, crear reporte vacío
    if not products:
        ws_resumen['A10'] = "⚠️ NO HAY PRODUCTOS EN EL INVENTARIO"
        ws_resumen['A10'].font = Font(name='Calibri', size=14, bold=True, color='FF0000')
        ws_resumen.merge_cells('A10:H10')
        ws_resumen['A10'].alignment = Alignment(horizontal='center')
        
        # Guardar reporte mínimo
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'inventario_vacio_joda_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        )
    
    total_products = len(products)
    total_value = sum(p.stock * p.cost_price for p in products)
    low_stock_products = [p for p in products if p.stock <= p.min_stock and p.stock > 0]
    no_stock_products = [p for p in products if p.stock == 0]
    
    # Estadísticas por categoría (solo categorías con productos)
    category_stats = {}
    for product in products:
        cat = product.category
        if cat not in category_stats:
            category_stats[cat] = {
                'productos': 0,
                'valor_inventario': 0,
                'stock_bajo': 0,
                'sin_stock': 0
            }
        category_stats[cat]['productos'] += 1
        category_stats[cat]['valor_inventario'] += product.stock * product.cost_price
        if product.stock <= product.min_stock and product.stock > 0:
            category_stats[cat]['stock_bajo'] += 1
        elif product.stock == 0:
            category_stats[cat]['sin_stock'] += 1
    
    # Métricas generales en tarjetas
    ws_resumen['A4'] = "📈 MÉTRICAS GENERALES"
    ws_resumen['A4'].font = subtitle_font
    ws_resumen.merge_cells('A4:H4')
    
    # Métricas
    ws_resumen['A6'] = "Total Productos"
    ws_resumen['B6'] = total_products
    ws_resumen['D6'] = "Valor Total Inventario"
    ws_resumen['E6'] = total_value
    ws_resumen['E6'].number_format = '$#,##0.00'
    
    ws_resumen['A8'] = "Productos Stock Bajo"
    ws_resumen['B8'] = len(low_stock_products)
    ws_resumen['D8'] = "Productos Sin Stock"
    ws_resumen['E8'] = len(no_stock_products)
    
    # Aplicar estilos a métricas
    for row in [6, 8]:
        for col in ['A', 'B', 'D', 'E']:
            cell = ws_resumen[f'{col}{row}']
            cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            cell.border = border_thin
            cell.font = Font(name='Calibri', size=11, bold=True)
            cell.alignment = Alignment(horizontal='center')
    
    # ============ HOJA 2: INVENTARIO POR CATEGORÍAS ============
    ws_inventario = wb.create_sheet("📦 Inventario por Categorías")
    
    # Título
    ws_inventario['A1'] = "📦 INVENTARIO DETALLADO POR CATEGORÍAS"
    ws_inventario['A1'].font = title_font
    ws_inventario.merge_cells('A1:I1')
    ws_inventario['A1'].alignment = Alignment(horizontal='center')
    
    current_row = 3
    
    # Solo procesar categorías que tienen productos reales
    if not category_stats:
        ws_inventario['A3'] = "⚠️ NO HAY PRODUCTOS PARA MOSTRAR"
        ws_inventario['A3'].font = Font(name='Calibri', size=14, bold=True, color='FF0000')
        ws_inventario.merge_cells('A3:I3')
        ws_inventario['A3'].alignment = Alignment(horizontal='center')
    else:
        # Procesar solo las categorías que tienen productos
        for category in sorted(category_stats.keys()):
            category_products = [p for p in products if p.category == category]
            if not category_products:
                continue
            
            # Encabezado de categoría
            ws_inventario[f'A{current_row}'] = f"🏷️ CATEGORÍA: {category.upper()}"
            ws_inventario[f'A{current_row}'].font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
            ws_inventario[f'A{current_row}'].fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
            ws_inventario.merge_cells(f'A{current_row}:I{current_row}')
            ws_inventario[f'A{current_row}'].alignment = Alignment(horizontal='center')
            current_row += 1
            
            # Headers de la tabla
            headers = ['Código', 'Producto', 'Stock Act.', 'Stock Mín.', 'Precio Costo', 'Precio Venta', 'Valor Inv.', 'Estado', 'Margen %']
            for col, header in enumerate(headers, 1):
                cell = ws_inventario.cell(row=current_row, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border_thin
                cell.alignment = Alignment(horizontal='center')
            current_row += 1
            
            # Datos de productos de la categoría
            for product in sorted(category_products, key=lambda x: x.name):
                # Color de fondo según categoría
                category_fill = PatternFill(start_color=CATEGORY_COLORS.get(category, 'FFFFFF'), 
                                          end_color=CATEGORY_COLORS.get(category, 'FFFFFF'), 
                                          fill_type="solid")
                
                # Color especial para stock bajo
                if product.stock <= product.min_stock and product.stock > 0:
                    row_fill = low_stock_fill
                elif product.stock == 0:
                    row_fill = no_stock_fill
                else:
                    row_fill = category_fill
                
                # Calcular margen
                margin = ((product.sell_price - product.cost_price) / product.cost_price * 100) if product.cost_price > 0 else 0
                
                # Insertar datos
                cells_data = [
                    (product.code, None),
                    (product.name, None),
                    (product.stock, None),
                    (product.min_stock, None),
                    (product.cost_price, '$#,##0.00'),
                    (product.sell_price, '$#,##0.00'),
                    (product.stock * product.cost_price, '$#,##0.00'),
                    (product.status.replace('_', ' ').title(), None),
                    (margin / 100, '0.0%')
                ]
                
                for col, (value, number_format) in enumerate(cells_data, 1):
                    cell = ws_inventario.cell(row=current_row, column=col, value=value)
                    cell.fill = row_fill
                    cell.font = data_font
                    cell.border = border_thin
                    if number_format:
                        cell.number_format = number_format
                    # Alineación especial para números
                    if col in [3, 4, 5, 6, 7, 9]:  # Columnas numéricas
                        cell.alignment = Alignment(horizontal='center')
                
                current_row += 1
            
            # Resumen de categoría
            ws_inventario[f'A{current_row}'] = f"📊 RESUMEN {category.upper()}"
            ws_inventario[f'A{current_row}'].font = Font(name='Calibri', size=11, bold=True)
            ws_inventario[f'A{current_row}'].fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            
            stats = category_stats[category]
            ws_inventario[f'C{current_row}'] = f"Productos: {stats['productos']}"
            ws_inventario[f'E{current_row}'] = f"Stock Bajo: {stats['stock_bajo']}"
            ws_inventario[f'G{current_row}'] = f"Valor: ${stats['valor_inventario']:,.2f}"
            
            for col in [1, 3, 5, 7]:
                ws_inventario.cell(row=current_row, column=col).fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
                ws_inventario.cell(row=current_row, column=col).border = border_thin
            
            current_row += 3  # Espaciado entre categorías
    
    # ============ HOJA 3: PRODUCTOS CON STOCK BAJO ============
    ws_stock_bajo = wb.create_sheet("⚠️ Stock Bajo")
    
    # Título
    ws_stock_bajo['A1'] = "⚠️ PRODUCTOS CON STOCK BAJO - ATENCIÓN REQUERIDA"
    ws_stock_bajo['A1'].font = Font(name='Calibri', size=16, bold=True, color='F57C00')
    ws_stock_bajo.merge_cells('A1:H1')
    ws_stock_bajo['A1'].alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['🔴', 'Código', 'Producto', 'Categoría', 'Stock Act.', 'Stock Mín.', 'Diferencia', 'Acción Sugerida']
    for col, header in enumerate(headers, 1):
        cell = ws_stock_bajo.cell(row=3, column=col, value=header)
        cell.fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
        cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        cell.border = border_thin
        cell.alignment = Alignment(horizontal='center')
    
    # Datos de productos con stock bajo (solo productos reales)
    all_low_stock = [p for p in products if p.stock <= p.min_stock]
    
    if not all_low_stock:
        ws_stock_bajo['A5'] = "✅ NO HAY PRODUCTOS CON STOCK BAJO"
        ws_stock_bajo['A5'].font = Font(name='Calibri', size=14, bold=True, color='00AA00')
        ws_stock_bajo.merge_cells('A5:H5')
        ws_stock_bajo['A5'].alignment = Alignment(horizontal='center')
        ws_stock_bajo['A6'] = "¡Felicidades! Todos los productos tienen stock adecuado."
        ws_stock_bajo['A6'].font = Font(name='Calibri', size=11, color='666666')
        ws_stock_bajo.merge_cells('A6:H6')
        ws_stock_bajo['A6'].alignment = Alignment(horizontal='center')
    else:
        for row, product in enumerate(sorted(all_low_stock, key=lambda x: x.stock), 4):
            # Prioridad visual
            if product.stock == 0:
                priority = "🔴"  # Crítico
                action = "REABASTECER URGENTE"
                row_color = "FFCDD2"
            elif product.stock <= product.min_stock * 0.5:
                priority = "🟠"  # Alto
                action = "Reabastecer pronto"
                row_color = "FFE0B2"
            else:
                priority = "🟡"  # Medio
                action = "Planificar reabastecimiento"
                row_color = "FFF9C4"
            
            fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
            
            data = [
                priority,
                product.code,
                product.name,
                product.category,
                product.stock,
                product.min_stock,
                product.min_stock - product.stock,
                action
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws_stock_bajo.cell(row=row, column=col, value=value)
                cell.fill = fill
                cell.font = data_font
                cell.border = border_thin
                if col in [5, 6, 7]:  # Columnas numéricas
                    cell.alignment = Alignment(horizontal='center')
    
    # ============ AJUSTAR ANCHOS DE COLUMNAS ============
    column_widths = {
        ws_resumen: [20, 15, 15, 20, 15, 15, 15, 15],
        ws_inventario: [12, 25, 10, 10, 12, 12, 12, 15, 10],
        ws_stock_bajo: [5, 12, 30, 15, 10, 10, 10, 25]
    }
    
    for worksheet, widths in column_widths.items():
        for i, width in enumerate(widths, 1):
            worksheet.column_dimensions[get_column_letter(i)].width = width
    
    # ============ GUARDAR Y RETORNAR ============
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'inventario_profesional_joda_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    )
